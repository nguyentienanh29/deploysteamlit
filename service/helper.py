import os
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

# --------- Excel ---------
def create_excel(file_excel: str, 
                 work_sheet_name:str, 
                 header_excel:str):
    
    if os.path.exists(file_excel):
        wb = load_workbook(file_excel)
        if work_sheet_name in wb.sheetnames:
            ws = wb[work_sheet_name]
        else:
            ws = wb.create_sheet(work_sheet_name)
            ws.append(header_excel)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = work_sheet_name
        ws.append(header_excel)   

    return wb, ws

def check_sheet_has_data(ws):
     for row in ws.iter_rows(min_row=2,values_only=True):
        for cell in row:
            if cell is not None:
                return True
     return False

def clear_sheet_data(ws):
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)


def pop_flash_message(session_state, key):
    if key in session_state:
        return session_state.pop(key)
    return None


def parse_int_input(raw_value, field_name):
    try:
        return int(str(raw_value).strip())
    except ValueError as exc:
        raise ValueError(f"{field_name} phải là số") from exc


def df_to_excel(df: pd.DataFrame):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="report")
    return output.getvalue()
